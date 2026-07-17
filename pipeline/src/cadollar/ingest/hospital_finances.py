"""Hospital annual financials (HCAI): the receiving end of Medi-Cal.

Every licensed California hospital files an annual financial disclosure with
HCAI. The Selected file carries net patient revenue BY PAYER — including
Medi-Cal managed care, the flow the site's dark-zone registry calls its
biggest gap. Plan-by-plan payments stay dark; each hospital's yearly Medi-Cal
total is public here, and that's what we publish.

Fail-honest: the latest fiscal year is largely "In Process" (unaudited); every
hospital row carries its report status, and the headline totals come from the
mostly-audited prior year, labeled as such.
"""

from __future__ import annotations

import hashlib
import io
import json
from datetime import UTC, datetime
from typing import Any

import openpyxl

from ..config import Settings
from ..publish.envelope import envelope
from ..sources import SourceConfig
from ..storage import Storage, read_manifest, write_manifest
from .csv_download import QualityGateError
from .http import fetch_bytes


def _num(v: Any) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _parse_year(blob: bytes) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(h) if h is not None else "" for h in next(rows)]
    out = []
    for r in rows:
        d = dict(zip(hdr, r))
        if not d.get("FAC_NO") or not d.get("FAC_NAME"):
            continue
        out.append(
            {
                "fac_no": str(d["FAC_NO"]),
                "name": str(d["FAC_NAME"]).strip(),
                "county": (d.get("COUNTY") or "").strip() or None,
                "control": (d.get("TYPE_CNTRL") or "").strip() or None,
                "care_type": (d.get("TYPE_CARE") or "").strip() or None,
                "report_status": (d.get("DATA_IND") or "").strip() or None,
                "net_patient_rev_usd": _num(d.get("NET_PT_REV")),
                "medical_ffs_usd": _num(d.get("NETRV_MCAL_TR")),
                "medical_managed_usd": _num(d.get("NETRV_MCAL_MC")),
                "medicare_usd": _num(d.get("NETRV_MCAR_TR")) + _num(d.get("NETRV_MCAR_MC")),
                "county_usd": _num(d.get("NETRV_CNTY")),
                "commercial_usd": _num(d.get("NETRV_THRD_TR")) + _num(d.get("NETRV_THRD_MC")),
                "operating_expenses_usd": _num(d.get("TOT_OP_EXP")),
                "salaries_usd": _num(d.get("EXP_SAL")) + _num(d.get("EXP_BEN")),
                "net_income_usd": _num(d.get("NET_INCOME")),
            }
        )
    wb.close()
    return out


def run_hospital_finances(storage: Storage, cfg: SourceConfig, settings: Settings) -> str:
    now = datetime.now(UTC)
    as_of = now.strftime("%Y-%m-%dT%H%M%SZ")
    manifest = read_manifest(storage, cfg.source)
    files: dict[str, str] = cfg.extra["files"]
    headline_fy: str = cfg.extra["headline_fy"]

    blobs = {fy: fetch_bytes(url) for fy, url in sorted(files.items())}
    content_hash = hashlib.sha256(b"".join(blobs[k] for k in sorted(blobs))).hexdigest()
    if manifest.get("content_hash") == content_hash:
        manifest["checked_at"] = now.isoformat()
        write_manifest(storage, cfg.source, manifest)
        print(f"{cfg.source}: unchanged, no-op")
        return manifest["published_key"]

    years: dict[str, dict[str, Any]] = {}
    hospitals: dict[str, dict[str, Any]] = {}
    total_rows = 0
    for fy, blob in blobs.items():
        rows = _parse_year(blob)
        total_rows += len(rows)
        audited = sum(1 for r in rows if r["report_status"] == "Audited")
        years[fy] = {
            "hospital_count": len(rows),
            "audited_count": audited,
            "in_process_count": len(rows) - audited,
            "net_patient_rev_usd": sum(r["net_patient_rev_usd"] for r in rows),
            "medical_ffs_usd": sum(r["medical_ffs_usd"] for r in rows),
            "medical_managed_usd": sum(r["medical_managed_usd"] for r in rows),
            "medicare_usd": sum(r["medicare_usd"] for r in rows),
        }
        for r in rows:
            h = hospitals.setdefault(
                r["fac_no"],
                {
                    "name": r["name"],
                    "county": r["county"],
                    "control": r["control"],
                    "care_type": r["care_type"],
                    "years": {},
                },
            )
            h["years"][fy] = {
                k: r[k]
                for k in (
                    "report_status", "net_patient_rev_usd", "medical_ffs_usd",
                    "medical_managed_usd", "medicare_usd", "county_usd",
                    "commercial_usd", "operating_expenses_usd", "salaries_usd",
                    "net_income_usd",
                )
            }
            # keep the freshest identity fields (names occasionally change)
            h["name"], h["county"] = r["name"], r["county"]

    if total_rows < cfg.min_rows:
        raise QualityGateError(f"{cfg.source}: only {total_rows} hospital-years parsed")

    # site join key: exact name (upper) -> fac_no
    name_index = {h["name"].upper(): fac for fac, h in hospitals.items()}
    doc = {
        "headline_fy": headline_fy,
        "years": years,
        "hospital_count": len(hospitals),
        "name_index": name_index,
        "hospitals": hospitals,
    }

    key = "published/hospital_finances.json"
    storage.put_bytes(
        key,
        json.dumps(
            envelope(cfg, as_of=as_of, ingested_at=now.isoformat(), data=doc), indent=2
        ).encode(),
        "application/json",
    )
    write_manifest(
        storage,
        cfg.source,
        {
            "content_hash": content_hash,
            "row_count": total_rows,
            "as_of": as_of,
            "ingested_at": now.isoformat(),
            "checked_at": now.isoformat(),
            "published_key": key,
        },
    )
    hl = years[headline_fy]
    print(
        f"{cfg.source}: {len(hospitals)} hospitals; {headline_fy} Medi-Cal net "
        f"${(hl['medical_ffs_usd'] + hl['medical_managed_usd']) / 1e9:.1f}B -> {key}"
    )
    return key
