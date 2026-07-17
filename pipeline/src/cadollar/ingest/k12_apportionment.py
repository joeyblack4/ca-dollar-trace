"""Principal Apportionment: what the state certifies it will pay each LEA.

Parses CDE's Principal Apportionment Summary workbook (one row per school
district, county office of education, or charter school) and publishes each
LEA's certified state aid, keyed by CDS code so the K-12 drill joins
deterministically to SACS spending — sent (this budget year) beside spent
(last completed year), never subtracted.
"""

from __future__ import annotations

import hashlib
import io
import json
import warnings
from datetime import UTC, datetime
from typing import Any

import openpyxl

from ..config import Settings
from ..publish.envelope import envelope
from ..sources import SourceConfig
from ..storage import Storage, read_manifest, write_manifest
from .csv_download import QualityGateError
from .http import fetch_bytes


def _find_col(hdr: list[str], needle: str) -> int:
    for i, h in enumerate(hdr):
        if h and needle in " ".join(str(h).split()):
            return i
    raise QualityGateError(f"k12_apportionment: column containing {needle!r} not found")


def _num(v: Any) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


def run_k12_apportionment(storage: Storage, cfg: SourceConfig, settings: Settings) -> str:
    now = datetime.now(UTC)
    as_of = now.strftime("%Y-%m-%dT%H%M%SZ")
    manifest = read_manifest(storage, cfg.source)

    blob = fetch_bytes(cfg.extra["file_url"])
    content_hash = hashlib.sha256(blob).hexdigest()
    if manifest.get("content_hash") == content_hash:
        manifest["checked_at"] = now.isoformat()
        write_manifest(storage, cfg.source, manifest)
        print(f"{cfg.source}: unchanged, no-op")
        return manifest["published_key"]

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")  # workbook carries odd print-area names
        wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    hdr: list[str] = []
    for r in rows:
        vals = [str(c) if c is not None else "" for c in r]
        if vals[0].strip() == "County Code":
            hdr = vals
            break
    if not hdr:
        raise QualityGateError(f"{cfg.source}: header row not found")

    i_total = _find_col(hdr, "Total Principal Apportionment")
    i_lcff_epa = _find_col(hdr, "Total of LCFF State Aid and EPA")
    i_epa = _find_col(hdr, "EPA Entitlement")
    i_sped = _find_col(hdr, "Special Education AB 602")

    leas: dict[str, dict[str, Any]] = {}
    for r in rows:
        cc = str(r[0] or "").strip()
        dc = str(r[1] or "").strip()
        sc = str(r[2] or "").strip()
        name = str(r[5] or "").strip()
        if not cc or not dc or not name or cc == "None":
            continue
        cds = f"{cc:0>2}{dc:0>5}{sc:0>7}"
        leas[cds] = {
            "name": name,
            "lea_type": str(r[6] or "").strip() or None,
            "is_charter": str(r[3] or "").strip() not in ("", "N/A", "None"),
            "total_apportionment_usd": round(_num(r[i_total])),
            "lcff_state_aid_epa_usd": round(_num(r[i_lcff_epa])),
            "epa_usd": round(_num(r[i_epa])),
            "special_ed_usd": round(_num(r[i_sped])),
        }
    wb.close()

    if len(leas) < cfg.min_rows:
        raise QualityGateError(f"{cfg.source}: only {len(leas)} LEAs parsed")

    doc = {
        "fiscal_year": cfg.extra["fiscal_year"],
        "certification": cfg.extra["certification"],
        "lea_count": len(leas),
        "charter_count": sum(1 for v in leas.values() if v["is_charter"]),
        "statewide_total_usd": sum(v["total_apportionment_usd"] for v in leas.values()),
        "statewide_special_ed_usd": sum(v["special_ed_usd"] for v in leas.values()),
        "leas": leas,
    }

    key = "published/k12_apportionment.json"
    storage.put_bytes(
        key,
        json.dumps(
            envelope(cfg, as_of=as_of, ingested_at=now.isoformat(), data=doc), indent=2
        ).encode(),
        "application/json",
    )
    write_manifest(
        storage,
        cfg.source,
        {
            "content_hash": content_hash,
            "row_count": len(leas),
            "as_of": as_of,
            "ingested_at": now.isoformat(),
            "checked_at": now.isoformat(),
            "published_key": key,
        },
    )
    print(
        f"{cfg.source}: {len(leas)} LEAs, ${doc['statewide_total_usd'] / 1e9:.1f}B certified "
        f"({doc['fiscal_year']} {doc['certification']}) -> {key}"
    )
    return key
